import pandas as pd
import openpyxl as op
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Protection, numbers, Font, PatternFill, Alignment, Border, Side
from io import BytesIO

class ExcelService:

    def download_excel_to_byte(df:pd.DataFrame, sheet_name:str):
        '''
            excel -> byte 변환
        '''

        # 0. 준비사항
        excel_byte:BytesIO = BytesIO()

        # 1. 변환
        with pd.ExcelWriter(excel_byte, engine='openpyxl') as writer:
            df.to_excel(writer, index=True, sheet_name=sheet_name)

            # 2. 작업
            # 2.0 준비사항
            wb:op.Workbook = writer.book # workbook
            ws:Worksheet = writer.sheets[sheet_name] # worksheet
            cell_range: str = ws.dimensions # 셀 범위
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)

            # 2.1 설정
            # 헤더 설정
            header_font = Font(name="맑은 고딕", # 폰트
                                size=12, # 크기
                                color="000000", # 글자 색상
                                bold=True, # 굴게
                                )
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 테두리 설정
            thin_gray_border = Border(left=Side(style='thin', color='4D4D4D'),
                                      right=Side(style='thin', color='4D4D4D'),
                                      top=Side(style='thin', color='4D4D4D'),
                                      bottom=Side(style='thin', color='4D4D4D')
                                      )
            # 열 너비 설정
            # index를 추가했으므로 +1 적용
            column_width_setting_info = [["주소", 40], ["도로명", 20], ["단지명", 20], ["거래금액(만원)", 15]]
            for info in column_width_setting_info:
                ws.column_dimensions[get_column_letter(ExcelService.get_column_index(df, info[0]) + 1)].width = info[1]

            # 2.2 적용
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    # 헤더 설정
                    if cell.row == min_row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # 전체 적용
                    cell.border = thin_gray_border

            # 틀고정
            ws.freeze_panes = "A2" # A1 행고정

        excel_byte.seek(0)

        return excel_byte
    
    def get_column_index(df: pd.DataFrame, column_name: str):
        '''
            컬럼 인덱스 조회
        '''

        # 1. df의 컬럼명 조회
        col_names = list(df.columns)

        # 2. df에서 찾고자하는 컬럼명 인덱스 조회
        col_index = col_names.index(column_name) + 1

        return col_index