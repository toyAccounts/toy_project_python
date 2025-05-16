from datetime import datetime
import pathlib
import openpyxl as op
from openpyxl.packaging.core import DocumentProperties
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.cell import Cell
from openpyxl.styles import Protection, numbers, Font, PatternFill, Alignment, Border, Side
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, LineChart, AreaChart, PieChart, RadarChart, BubbleChart, Reference
from openpyxl.chart.series import Series
from openpyxl.chart.marker import DataPoint
import csv
from pprint import pprint
from operator import itemgetter
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows

################################################## chapter 4. ##################################################
# 경로
path = pathlib.Path(r".\data")

# 경로 내 엑셀파일 조회
# for path_obj in path.iterdir():
#     if(path_obj.match("*.xlsx")):
for path_obj in path.glob("*.xlsx"):

    # workbook
    wb:op.Workbook = op.load_workbook(path_obj)

    # 시트 이름 목록
    print("시트 목록:", wb.sheetnames)

    # 시트 활성화 조회
    ws:Worksheet = wb.active
    print("활성 시트 이름:", ws.title)

    # 시트 확인
    for sheet_name in wb.sheetnames:

        # 시트 선택
        # 방법 1)
        ws:Worksheet = wb[sheet_name]

        # 방법 2)
        # ws = wb.worksheets[0]

        print(f"\n시트: {sheet_name} / title: {ws.title}")
        print(f' - 시트 상태: {ws.sheet_state}')
        print(" - 데이터 범위:", ws.dimensions)  # 예: A1:E15
        print(" - 최대 행:", ws.max_row)
        print(" - 최대 열:", ws.max_column)

    # 활성 시트 변경
    if(ws.title == "Sheet5"):
        index:int = wb.sheetnames.index("Sheet5")
    # wb.active = index

    # 속성
    wb_property:DocumentProperties = wb.properties
    # wb_property.creator = "작성자"
    # wb_property.title = "제목"
    # wb_property.last_modified_by = "최근 수정자"

    # 특정 셀 값
    print(f'A1: {ws["A1"].value}')

    # wooksheet 보호
    wb.security = WorkbookProtection(workbookPassword="test",
                                    #  lockStructure=True
                                    lockStructure=False
                                    )

    ################################################## chapter 5. ##################################################
    # 시트 생성
    # wb_new_sheet = wb.create_sheet(title="새로운 시트")

    # 시트명 변경
    # ws = wb["Sheet1"]
    # ws.title = "표222"

    # 시트 삭제
    # wb.remove(ws)


    # 시트 조작
    for ws in wb:
        # ws["B2"].value = ws.title

        # 시트 이동시 그룹화 해제
        ws.sheet_view.tabSelected = None



    # 시트 이동
    # offset: 이동하는 수 (양수: 오른쪽 이동 / 음수: 왼쪽 이동)
    # active한 시트 이동시 그룹화 됨
    if(ws.title == "새로운 시트"):
        wb.move_sheet("", offset=-3)


    # 시트 복사
    ws_copy_target:Worksheet = wb.worksheets[0]
    ws_copy = wb.copy_worksheet(ws_copy_target)
    ws_copy.title =f'{ws_copy_target.title}의 복사본'


    # 시트 표시 / 비표시
    ws_copy_target.sheet_state = Worksheet.SHEETSTATE_HIDDEN

    # 시트 보호
    # 입력가능한 셀의 잠금 off -> 시트의 보호 on
    ws_protect_target:Worksheet = wb.worksheets[0]

    for row in ws_protect_target["B2:D6"]:
        
        for cell in row:
            # 1) 셀의 잠금 off
            # cell_protect_target:Cell = ws_protect_target["B2"]
            # cell_protect_target.protection = Protection(locked=False)
            cell.protection = Protection(locked=False)

    # 2) 시트 보호 on
    ws_protect_target.protection.password = "sheet"
    ws_protect_target.protection.enable()



    ################################################## chapter 6. ##################################################

    # 행 추가 (수식범위가 연관된 경우 범위자동수정 안됨)
    ws:Worksheet = wb.active
    print(f'활성화된 시트명: {ws.title}')
    for i, v in enumerate(["D1", "D2", "D3"]):
        cell:Cell = ws[v]
        cell.value = i

    ws["D4"] = "=sum(D1:D3)"
    ws.insert_rows(2) # 2번째 행에 추가

    ws["D5"] = "=sum(D1:D4)" # 관련 수식 수정


    # 행 삭제
    ws.delete_rows(10)



    # 열 추가
    ws.insert_cols(3)

    for row in ws.iter_rows(min_row=2):
        cell:Cell = row[4] # 4 == D
        cell.value = f'=B{row[0].row}*E{row[0].row}' # 수식처리
        # cell.value = row[1].value * row[3].value # 직접 계산

    # 열 삭제
    ws.delete_cols(3)

    # 마지막 열, 행
    print(f'마지막 행: {ws.max_row}, 마지막 열; {ws.max_column}')
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell: Cell = ws.cell(row, col)
            print(cell.value)




    # 행 높이 맞추기
    # 행번호
    ws.row_dimensions[1].height = 240

    # 열 폭 맞추기
    # 열 이름
    ws.column_dimensions["A"].width = 88

    # 틀고정
    ws.freeze_panes = "B2" # A1 고정
    # ws.freeze_panes = "A2" # 행고정 (A2, A3, ...)
    # ws.freeze_panes = "B1" # 열고정 (B1, C1, ...)


    # 틀고정 해제
    # ws.freeze_panes = None


    # 행 숨김/표시
    ws.row_dimensions[3].hidden = True

    # 열 숨김/표시
    ws.column_dimensions["A"].hidden = True

    # 열 이름 확인
    for col_index in range(3, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_index).column_letter # 열 이름
        print(f'col_letter: {col_letter}')
        ws.column_dimensions[col_letter].hidden = False


    ################################################## chapter 7. ##################################################
    # 셀 값
    # 방법 1)
    cell:Cell = ws["C2"]
    cell.value

    # 방법 2)
    # row, column은 모두 1부터 시작
    cell:Cell = ws.cell(row=2, column=3)
    cell.value

    # 셀값 추가
    cell.value = "값"


    # 셀값 범위
    # 방법 1)
    cell_scope:tuple[tuple[Cell, ...], ...] = ws["B2:C6"]

    for row in cell_scope:
        for cell in row:
            print(cell)
    print("-" * 50)

    # 방법 2)
    for row in ws.iter_rows(min_row=2, max_row=6,
                            min_col=2, max_col=3):
        for cell in row:
            print(cell)
    print("-" * 50)
    
    # 방법 3)
    for row_index in range(2, 7):
        for col_index in range(2, 4):
            print(ws.cell(row=row_index, column=col_index))

    print(type(cell_scope))


    # 셀 범위 값 -> 다른 셀 범위에 복사
    mov_row = 5
    mov_col = 4
    for row_index in range(2, 7):
        for col_index in range(2, 4):
            ws.cell(row=(row_index + mov_row),
                    column=(col_index + mov_col)).value = ws.cell(row=row_index, column=col_index).value

    # 행렬 변경
    mov_row = 10
    for row_index in range(2, 7):
        for col_index in range(2, 4):
            ws.cell(row=(col_index + mov_row),
                    column=(row_index)).value = ws.cell(row=row_index, column=col_index).value


    # 다른 시트에 옮겨 적기
    if(len(wb.worksheets) > 3):
        ws_3 = wb.worksheets[3]
        ws_4 = wb.worksheets[4]

        for row in cell_scope:
            for cell in row:
                ws_3.cell(row_index, col_index).value = ws_4.cell(row_index, col_index).value

        # 다른 북에 옮겨 적기
        ws_5 = wb.worksheets[5] # 복사 대상
        wb_new_1 = op.Workbook() # 붙이기 대상
        ws_new_1 = wb_new_1.active

        for row in cell_scope:
            for cell in row:
                ws_new_1.cell(row_index, col_index).value = ws_5.cell(row_index, col_index).value


    print(f"stem: {path_obj.stem} / suffix :{path_obj.suffix}")
    print(type(path_obj)) 
    wb_new_1.save(path_obj.with_name(f'{path_obj.stem}_copy{path_obj.suffix}'))


    ################################################## chapter 8. ##################################################
    # 수치 표시 방식 설정
    '''
        #,##0 : 3자릿수마다 제로 억제 (3자리마다 ,을 넣음 + 맨 앞의 0을 제거)
        #,##0.00: 소수점 이하 자리수 지정
    '''
    ws.cell(1, 2).value = 12345678.987654321
    # ws.cell(1, 2).number_format = "#,##0.0"
    ws.cell(1, 2).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1 # '#,##0.00'


    # 날짜 서식
    cell:Cell = ws.cell(1, 5)
    cell.value = datetime(2025, 5, 12)
    # cell.number_format = "yyyy-mm-dd"
    cell.number_format = "yyyy년 mm월 dd일"
    # cell.number_format = numbers.FORMAT_DATE_DMYSLASH

    # 글자 설정
    font_01 = Font(name="고딕", # 폰트
                       size=20, # 크기
                       color="000FFF", # 문자 색ㅇ
                       bold=True, # 굴게
                       italic=True, # 이텔릭체
                       underline="single", # 밑줄
                       strike=True # 취소선
                    )
    
    # 셀 채우기
    fill_01 = PatternFill(patternType="darkVertical", # 채우기 패턴
                          fgColor="A3E312", # 채우기 색상
                          bgColor="FFFF00" # 패턴 배경 (패턴이 solid인 경우 적용x)
                          )
    
    # 문자 배치 지정
    al_01 = Alignment(horizontal="left", # 가로 위치 (left: 왼쪽 맞춤, center: 가운데 맞춤, right: 오른쪽 맞춤, distributed: 균등 분할)
                      vertical="bottom" # 세로 위치 (top: 위쪽 맞춤, center: 위아래쪽 가운데 맞춤, bottom: 아래쪽 맞춤)
                      )

    # 괘선 긋기
    '''
        스타일
            - 실선: hair: 가장 가늚, thin: 가늚, medium: 중간, thick: 두꺼움
            - 그외: dashed: 파선, dotted: 점선, dashDot: 일점 쇄선, dashDotDot: 이점 쇄선, double: 이중선
    '''
    side_01 = Side(style="thin", # 선 스타일
                   color="000000" # 선 색
                   )
    border_01 = Border(left=side_01,
                    #    right=side_01,
                       top=side_01,
                       bottom=side_01)
    
    # 적용
    for rows in ws["A1:F3"]:
        # rows = tuple(Cell)
        first_cell:Cell = rows[0]
        row_index = first_cell.row  # 첫 번째 셀의 행 번호 추출
        print(f'row_index: {row_index}')
        ws.row_dimensions[row_index].height = (30  + (5 * row_index))
    
        for cell in rows:
            cell.font = font_01 # 글자
            cell.fill = fill_01 # 채우기
            cell.alignment = al_01 # 문자 배치
            cell.border = border_01 # 괘선


    # 셀 결합
    ws.merge_cells("K4:P8")
    ws["K4"].value = "셀 결합"

    # 셀 결합 해제
    # ws.unmerge_cells("K4:P8")
    

    ################################################## chapter 9. ##################################################
    # 셀 수식 지정
    # 수식은 문자열로 다룸
    ws["C35"] = "=A15+B15"

    for row in ws.iter_rows(min_row=20, max_row=25):
        row[1].value = f"=C{row[0].row}*D{row[0].row}"

    # 셀에 엑셀 함수 설정
    for rows in ws["F9:F18"]:
        for cell in rows:
            cell.value = 0.3

    ws["F19"] = "=SUM(F9:F18)"
    ws["F20"] = "=ROUNDDOWN(F19 * 0.1, 0)"


    # 함수식 상대 참조 복사
    # translate_formula: 상대 참조로 복사
    # F19에 지정된 수식 (sum(f9:f18))을 D19에 적용 (sum(d9:d18))
    ws["D19"] = Translator(ws["F19"].value, # 상대 참조
                           origin="F19" # 상대 참조의 기점으로 하는 셀번지
                           ).translate_formula("D19") # 복사 대상
    


    # 데이터 입력 규칙
    '''
        - 입력값 타입
            - whole: 정수값
            - decimal: 소수값
            - date: 날짜
            - time: 시간
            - textLength: 문자열의 길이
            - list: 리스트
            - Custom: 커스텀
    '''
    dv_01 = DataValidation(type="list", # 입력값 타입
                           formula1='"완납, 일부 납품"', # 입력할 수 있는 값
                           allow_blank=True # 공백허용 여부
                           )
    # 규칙 적용 범위 설정
    dv_01.add("O9:O18")

    # 규칙 적용
    ws.add_data_validation(dv_01)


    # 조건부 서식
    less_than_rule = CellIsRule(operator="lessThan",
                                formula=[100],
                                stopIfTrue=True,
                                fill=PatternFill("solid", start_color="FF0000", end_color="0000FF"))
    
    two_color_scale = ColorScaleRule(start_type="min", start_color="FF0000",
                                     end_type="max", end_color="FFFFFF")
    # 조건부 서식 적용
    ws.conditional_formatting.add("B2:B100", less_than_rule)
    ws.conditional_formatting.add("C2:C100", two_color_scale)


    ################################################## chapter 10. ##################################################
    
    ws = wb["Sheet2"]

    # 막대 그래프 ####################################################################################################
    # y축 표시 안됨
    data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)
    
    # 차트 생성
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "과일 소비량"
    chart.x_axis.title = "시"
    chart.y_axis.title = "소비량"

    chart.width = 15 # 차트 폭
    chart.height = 10 # 차트 높이

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "H2")
    print(f'max low :{ws.max_row}')



    # 누적 막대 그래프 ####################################################################################################
    data = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)

    # 차트 생성
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "percentStacked"
    chart.overlap = 100
    chart.title = "과일 소비량"
    chart.x_axis.title = "시"
    chart.y_axis.title = "과일"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "H20")


    # 꺽은선 그래프 ####################################################################################################
    data = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)

    # 차트 생성
    chart = LineChart()
    chart.title = "과일 소비량"
    chart.x_axis.title = "시"
    chart.y_axis.title = "과일"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "H35")

    # 영역형 차트 ####################################################################################################
    data = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)

    # 차트 생성
    chart = AreaChart()
    chart.grouping = "stacked"
    chart.title = "시별 과일 소비량"
    chart.x_axis.title = "시"
    chart.y_axis.title = "과일"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "H55")

    # 파이 차트 ####################################################################################################
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    # 차트 생성
    chart = PieChart()
    chart.title = "시의 포도 소비량"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    # 차트 자르기
    slice:DataPoint = DataPoint(idx=0, # 잘라내려는 index
                                explosion=30 # 원그래프와 벌어진 정도
                                )
    series_obj:Series = chart.series[0]
    series_obj.data_points = [slice]

    ws.add_chart(chart, "H75")


    # 레이더 차트 ####################################################################################################
    data = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)

    # 차트 생성
    chart = RadarChart()
    chart.title = "과일의 소비량"
    chart.type = "filled"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "H95")


    # 버블 차트 ####################################################################################################
    # 차트 생성
    # 안됨
    '''
    chart = BubbleChart()
    chart.style = 18
    for row in range(2, ws.max_row + 1):
        x_values = Reference(ws, min_col=3, min_row=row)
        y_values = Reference(ws, min_col=2, min_row=row)
        size = Reference(ws, min_col=4, min_row=row)
        series = Series(valuse=y_values, xvalues=x_values, zvalues=size, title=ws.cell(row, 1).value)
        chart.series.append(series)
    ws.add_chart(chart, "H115")
    '''

    ################################################## chapter 11. ##################################################
    ws = wb["Sheet3"]

    ##### csv 파일 저장 ##################################################
    with open(r".\data\해테스트.csv", "w", encoding="utf_8_sig") as file_pointer:
        writer = csv.writer(file_pointer, lineterminator="\n")

        for row in ws.rows:
            # values = []
            # for cell in row:
            #     values.append(cell.value)
            # writer.writerow(values)
            writer.writerow([cell.value for cell in row])

    ##### 테이블 서식 지정 ##################################################
    table = Table(displayName="Table1",
                  ref="A1:F6")
    table_style = TableStyleInfo(name="TableStyleMedium2",
                                 showRowStripes=True # 행마다 패턴 설정
                                 )
    
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # 페이지 나누기
    for i in range(1, 3):
        page_break = Break(i * 2)
        ws.row_breaks.append(page_break)


    ##### 인쇄 설정 ##################################################
    '''
        &D: 현재 날짜
        &F: 파일 명
        &A: 시트 표제
        &T: 인쇄 시각
        &P: 페이지 번호
        &N: 총 페이지 수
    '''
    # 머리글 설정
    # ws.oddHeader.left.text = "&D"
    # ws.oddHeader.center.text = "&F"
    # ws.oddHeader.right.text = "&A"

    # 바닥글 설정
    # ws.oddFooter.center.text = "&P / &N페이지"

    # 표제 행 설정
    # 안됨
    ws.print_title_rows = "1:1" # 1행~1행까지 타이틀행으로 각 페이지에 인쇄됨

    # 인쇄 방향 설정
    # landscape: 가로방향, portrait: 세로방향
    # 전체의 열을 1페이지로 인쇄
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1 # 전체 열을 1페이지로 인쇄
    ws.page_setup.fitToHeight = 0 # 세로 지정x
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 용지 크기
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # 특정 범위만 중앙에 위치
    ws.print_area = "A1:F6" # 인쇄범위 지정
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True




    ################################################## chapter 12. ##################################################
    ws = wb["Sheet4"]

    ##### 추출 ##################################################
    # 방법 1)
    # for row in range(2 , ws.max_row + 1):
    #     cell_d:Cell = ws[f'I{row}']
    #     cell_h:Cell = ws[f'H{row}']
    #     value_d_slice: str = cell_d.value
    #     if(value_d_slice == "볼펜" or
    #        value_d_slice == "연필" or
    #        value_d_slice == "노트"):
    #         print(f'{cell_h.value}: {value_d_slice}')

    # 방법 2)
    # for row in range(2 , ws.max_row + 1):
    #     cell_d:Cell = ws[f'D{row}']
    #     value_d_slice: str = cell_d.value[:2]
    #     if(value_d_slice == "아림"):
    #         print(f'{value_d_slice}')

    ##### 별도 저장 ##################################################
    # wb_out = op.Workbook()
    # ws_out = wb_out.active
    # list_row = 1
    # for row in ws.iter_rows():
    #     if row[8].value[:2] == "연필" or list_row == 1:
    #         for cell in row:
    #             # 날짜 서식
    #             if (cell.col_idx == 2 and list_row != 1):
    #                 ws_out.cell(list_row, cell.col_idx).value = cell.value.date()
    #             else:
    #                 ws_out.cell(list_row, cell.col_idx).value = cell.value

    #         list_row += 1
    # wb_out.save(path_obj.with_name(f'{path_obj.stem}_추출_{path_obj.suffix}'))


    ##### 정렬 ##################################################
    sales_list:list[dict] = []
    for row in ws.iter_rows():
        print("-" * 50)
        if row[0].row == 1:
            header_cell = row
        else:
            row_dic:dict = {}

            for k, v in zip(header_cell, row):
                if(v.value == None):
                    print(f'k: {k.value} / v: {v.value}')
                row_dic[k.value] = v.value if v.value != None else "C005"
            
            sales_list.append(row_dic)

    # pprint(sales_list, sort_dicts=False)
    sorted_list_a = sorted(sales_list, key=itemgetter("거래처코드", "날짜"))
    # pprint(sorted_list_a, sort_dicts=False)
    # sales_list.sort(reverse=True)
    # pprint(sales_list, sort_dicts=False)

    ##### 집계 ##################################################
    # 1)
    sorted_list_a = sorted(sales_list, key=itemgetter("상품코드", "날짜"))
    wb2 = op.Workbook()
    ws2:Worksheet = wb2.active
    ws2.title = "상품 코드별 수량"
    list_row = 1

    ws2.cell(list_row, 1).value = "상품코드"
    ws2.cell(list_row, 2).value = "날짜"
    ws2.cell(list_row, 3).value = "품명"
    ws2.cell(list_row, 4).value = "수량"
    ws2.cell(list_row, 5).value = "합계"

    sum_q = 0
    old_key = ""
    for dic in sorted_list_a:
        # print(f'old_key: {old_key}')
        if old_key == "":
            old_key = dic["상품코드"]
        
        if old_key == dic["상품코드"]:
            sum_q += dic["수량"]
        else:
            ws2.cell(list_row, 5).value = sum_q
            sum_q = dic["수량"]
            old_key = dic["상품코드"]

        list_row += 1

        ws2.cell(list_row, 1).value = dic["상품코드"]
        ws2.cell(list_row, 2).value = dic["날짜"].date()
        ws2.cell(list_row, 3).value = dic["품명"]
        ws2.cell(list_row, 4).value = dic["수량"]

    ws2.cell(list_row, 5).value = sum_q
    wb2.save(path_obj.with_name(f'{path_obj.stem}_집계_{path_obj.suffix}'))

    # 2)
    sorted_list_b = sorted(sales_list, key=itemgetter("거래처코드", "상품코드", "날짜"))
    wb3 = op.Workbook()
    ws3:Worksheet = wb3.active





    # 시트 분리
    old_key = ""
    for dic in sorted_list_b:
        if old_key != dic["거래처코드"]:
            old_key = dic["거래처코드"]
            ws3 = wb3.create_sheet(title=dic["거래처명"])

            list_row = 1

            ws3.cell(list_row, 1).value = "상품코드"
            ws3.cell(list_row, 2).value = "날짜"
            ws3.cell(list_row, 3).value = "품명"
            ws3.cell(list_row, 4).value = "수량"
            ws3.cell(list_row, 5).value = "합계"

        list_row += 1

        ws3.cell(list_row, 1).value = dic["상품코드"]
        ws3.cell(list_row, 2).value = dic["날짜"].date()
        ws3.cell(list_row, 3).value = dic["품명"]
        ws3.cell(list_row, 4).value = dic["수량"]

    # wb3.remove(ws3["Sheet"])

    # 각 시트 내용 작업
    for ws3 in wb3:
        sum_q = 0
        old_key = ""
        for i in range(2, ws3.max_row + 1):
            print(f'old_key: {old_key}')
            if old_key == "":
                old_key = ws3.cell(i, 1).value

            if old_key == ws3.cell(i, 1).value:
                sum_q += ws3.cell(i, 4).value
            else:
                ws3.cell(i-1, 5).value = sum_q
                sum_q = ws3.cell(i, 4).value
                old_key = ws3.cell(i, 1).value
        ws3.cell(i, 5).value = sum_q

    wb3.save(path_obj.with_name(f'{path_obj.stem}_회사별_집계_{path_obj.suffix}'))

    ##### 집계(dict 사용) ##################################################
    ws = wb["Sheet4"]
    def print_header(ws: Worksheet):
        ws["A1"].value = "담당자"
        ws["B1"].value = "금액"
        ws["C1"].value = "수량"
        ws["D1"].value = "거래처"
        ws["E1"].value = "수량"
        ws["F1"].value = "금액"

    sales_data:dict = {}

    for row in range(2, ws.max_row + 1):
        person = ws[f"E{row}"].value
        customer = ws[f"C{row}"].value
        quantity = ws[f"J{row}"].value
        amount = ws[f"L{row}"].value

        print(f'person :{ person}  / customer: {customer}  / quantity:{quantity} / amount: {amount}')


        sales_data.setdefault(person, {"name": ws[f"F{row}"].value,
                                       "quantity": 0,
                                       "amount": 0})
        
        sales_data[person].setdefault(customer, {"name": ws[f"D{row}"].value,
                                       "quantity": 0,
                                       "amount": 0})

        sales_data[person][customer]["quantity"] += int(quantity)
        sales_data[person][customer]["amount"] += int(amount)
        sales_data[person]["quantity"] += int(quantity)
        sales_data[person]["amount"] += int(amount)

    wb4 = op.Workbook()
    ws4:Worksheet = wb4.active
    print_header(ws4)
    row = 2
    for person_data in sales_data.values():
        ws4[f'A{row}'].value = person_data["name"]
        ws4[f'B{row}'].value = person_data["quantity"]
        ws4[f'C{row}'].value = person_data["amount"]

        for customer_data in person_data.values():
            if isinstance(customer_data, dict):
                for item in customer_data.values():
                    ws4[f'D{row}'].value = customer_data["name"]
                    ws4[f'E{row}'].value = customer_data["quantity"]
                    ws4[f'F{row}'].value = customer_data["amount"]
                row += 1

    ws4[f'F{row}'].value = f"=SUM(F2:F{row-1})"
    ws4[f'E{row}'].value = "합계"

    # 저장
    wb4.save(path_obj.with_name(f'{path_obj.stem}_담당자_거래처별_집계_{path_obj.suffix}'))



    ##### 크로스 집계(list 사용) ##################################################
    customers:list = []
    products:list = []

    ws_in1 = wb["거래처"]
    for row in range(1, ws_in1.max_row + 1):
        customer = [ws_in1[f'A{row}'].value, ws_in1[f'B{row}'].value]
        customers.append(customer)

    ws_in2 = wb["상품"]
    for row in range(1, ws_in2.max_row + 1):
        product = f'{ws_in2[f'A{row}'].value}:{ws_in2[f'B{row}'].value}'
        products.append(product)


    sales_amount = [ [0] * (len(products) + 2) for i in range(len(customers) + 1)]


    for j in range(2, len(products) + 2):
        sales_amount[0][j] = products[j - 2]

    sales_amount[0][0] = "거래처코드"
    sales_amount[0][1] = "거래처명"

    for i in range(1, len(customers) + 1):
        sales_amount[i][0] = customers[i-1][0]
        sales_amount[i][1] = customers[i-1][1]

    ws_in3 = wb["매출명세"]
    for row in range(2, ws_in3.max_row + 1):
        customer = ws_in3[f"C{row}"].value
        product = ws_in3[f"H{row}"].value
        amount = ws_in3[f"J{row}"].value

        for i in range(len(customers) + 1):
            if(customer == sales_amount[i][0]):
                for j in range(2, len(products) + 2):
                    if(product == sales_amount[0][j].split(":")[0]):
                        sales_amount[i][j] += amount

    wb5 = op.Workbook()
    ws5:Worksheet = wb5.active
    row = 1

    print(customers)
    print("=" * 100)
    print(products)
    print("=" * 100)
    print(sales_amount)

    for sales_row in sales_amount:
        col = 1
        customer_sum = 0
        for sales_col in sales_row:
            print(f'sales_col: {sales_col}')
            if row == 1 and col > 2:
                ws5.cell(row, col).value = sales_col.split(":")[1]
            else:
                ws5.cell(row, col).value = sales_col
            

            if(row > 1 and col > 2):
                customer_sum += sales_col
            
            col += 1

        if(row == 1):
            ws5.cell(row, col).value = "합계"
        else:
            ws5.cell(row, col).value = customer_sum
        row += 1


    # 저장
    wb5.save(path_obj.with_name(f'{path_obj.stem}_크로스_집계_{path_obj.suffix}'))


    ##### pandas 사용 ##################################################
    data_folder_path:str = f'{os.path.dirname(os.path.abspath(__file__))}\\data' # 데이터 폴더 경로
    file_path = os.path.join(data_folder_path, "통합 문서1.xlsx")
    df = pd.read_excel(file_path, sheet_name="매출명세")


    print(df)
    print(f'size: {df.size} / shape :{df.shape} / len: {len(df)}')
    

    df2 = df.pivot_table(index="거래처코드", # 세로 집계항목
                         columns="상품코드", # 가로 집계항목
                         values="수량", # 집계 대상의 값 항목
                         fill_value=0, # NaN (결손값)을 채우는 값
                         margins=True,
                         aggfunc="sum" # 집계 함수
                         )
    
    print(df2)

    save_file_path = os.path.join(data_folder_path, "통합 문서1_pandas.xlsx")
    with pd.ExcelWriter(save_file_path) as writer:
        df2.to_excel(writer, sheet_name="거래처 상품별 수량")





    wb6 = op.Workbook()
    ws6:Worksheet = wb6.active

    for row in dataframe_to_rows(df2, index=True, header=True):
        ws6.append(row)


    ws6["A2"].value = "거래처명"


    ws_in1 = wb["거래처"]
    for row in range(2, ws6.max_row + 1):
        customer = ws6[f'A{row}'].value
        
        for row_in in range(1, ws_in1.max_row + 1):
            customer_in = ws_in1[f'A{row_in}'].value

            if(customer == customer_in):
                ws6[f'A{row}'].value = ws_in1[f'B{row_in}'].value

    ws_in2 = wb["상품"]
    for col in range(2, ws6.max_column + 1):
        product = ws6.cell(1, col).value
        for row_in in range(1, ws_in2.max_row + 1):
            product_in = ws_in2[f'A{row_in}'].value
            if(product == product_in):
               ws6.cell(1, col).value = ws_in2[f'B{row_in}'].value



    df3 = df[['거래처명', '품명', '금액']].query("금액 > 50000")
    save_file_path = os.path.join(data_folder_path, "통합 문서1_filter_pandas.xlsx")
    df3.to_excel(save_file_path)

    df4 = df.groupby(["담당자명", "거래처명"])["금액"].sum().reset_index()
    save_file_path = os.path.join(data_folder_path, "통합 문서1_groupby_pandas.xlsx")
    df4.to_excel(save_file_path)

    # 저장
    wb6.save(path_obj.with_name(f'{path_obj.stem}_openpyxl_pandas_{path_obj.suffix}'))



